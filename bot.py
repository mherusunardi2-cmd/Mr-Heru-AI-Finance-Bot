import telebot
from telebot import types
import sqlite3
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ==================== CONFIGURATION (CLOUD SECURE) ====================
# Mengambil token dari sistem hosting (Environment Variable) untuk keamanan di GitHub
# Jika dijalankan lokal di VS Code dan token sistem kosong, ia akan memakai token cadangan di bawah
BOT_TOKEN = os.environ.get("BOT_TOKEN", "8729434907:AAHEAaKSpAM23vTsK9l59_uqaVWLLlInGls")
DB_NAME = "keuangan.db"
bot = telebot.TeleBot(BOT_TOKEN)

user_state = {}

# ==================== DATABASE SETUP ====================
def init_db():
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS transaksi (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tanggal TEXT,
            jenis TEXT,
            kategori TEXT,
            nominal REAL,
            keterangan TEXT
        )
    ''')
    conn.commit()
    conn.close()

init_db()

# ==================== HANDLERS ====================
@bot.message_handler(commands=['start', 'menu'])
def send_welcome(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.row("📥 Catat Keuangan", "📊 Rekap & Export Excel")
    bot.send_message(
        message.chat.id, 
        f"Halo! Saya Akuntan Pribadi Anda.\nSilakan pilih menu di bawah ini untuk memulai:", 
        reply_markup=markup
    )

@bot.message_handler(func=lambda message: message.text == "📥 Catat Keuangan")
def start_pencatatan(message):
    markup = types.InlineKeyboardMarkup()
    markup.add(
        types.InlineKeyboardButton("Pemasukan (Income)", callback_data="jenis_Pemasukan"),
        types.InlineKeyboardButton("Pengeluaran (Expense)", callback_data="jenis_Pengeluaran")
    )
    bot.send_message(message.chat.id, "Pilih jenis transaksi:", reply_markup=markup)

@bot.callback_query_handler(func=lambda call: call.data.startswith("jenis_"))
def handle_jenis(call):
    jenis = call.data.split("_")[1]
    user_state[call.message.chat.id] = {"jenis": jenis}
    
    markup = types.InlineKeyboardMarkup()
    if jenis == "Pemasukan":
        kategori = ["Gaji", "Investasi", "Project", "Lainnya"]
    else:
        kategori = ["Makanan", "Transportasi", "Kontrakan", "Belanja", "Hiburan", "Lainnya"]
        
    for kat in kategori:
        markup.add(types.InlineKeyboardButton(kat, callback_data=f"kat_{kat}"))
        
    bot.edit_message_text(f"Jenis: {jenis}\nPilih Kategori:", call.message.chat.id, call.message.message_id, reply_markup=markup)

@bot.callback_query_handler(func=lambda call: call.data.startswith("kat_"))
def handle_kategori(call):
    kategori = call.data.split("_")[1]
    chat_id = call.message.chat.id
    user_state[chat_id]["kategori"] = kategori
    
    bot.delete_message(chat_id, call.message.message_id)
    msg = bot.send_message(chat_id, f"Jenis: {user_state[chat_id]['jenis']}\nKategori: {kategori}\n\nMasukkan NOMINAL (Angka saja, misal: 50000):")
    bot.register_next_step_handler(msg, process_nominal)

def process_nominal(message):
    chat_id = message.chat.id
    try:
        nominal = float(message.text.replace(".", "").replace(",", ""))
        user_state[chat_id]["nominal"] = nominal
        msg = bot.send_message(chat_id, "Masukkan KETERANGAN singkat:")
        bot.register_next_step_handler(msg, process_keterangan)
    except ValueError:
        msg = bot.send_message(chat_id, "❌ Input salah. Masukkan angka saja tanpa titik/koma:")
        bot.register_next_step_handler(msg, process_nominal)

def process_keterangan(message):
    chat_id = message.chat.id
    keterangan = message.text
    user_state[chat_id]["keterangan"] = keterangan
    
    data = user_state[chat_id]
    tanggal = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO transaksi (tanggal, jenis, kategori, nominal, keterangan) VALUES (?, ?, ?, ?, ?)",
        (tanggal, data['jenis'], data['kategori'], data['nominal'], keterangan)
    )
    conn.commit()
    conn.close()
    
    bot.send_message(
        chat_id, 
        f"✅ **Data Berhasil Disimpan!**\n\n📅 Tanggal: {tanggal}\n🗂 Jenis: {data['jenis']}\n🏷 Kategori: {data['kategori']}\n💰 Nominal: Rp {data['nominal']:,.0f}\n📝 Ket: {keterangan}"
    )
    user_state.pop(chat_id, None)

# ==================== EXCEL GENERATION ENGINE ====================
@bot.message_handler(func=lambda message: message.text == "📊 Rekap & Export Excel")
def handle_report(message):
    chat_id = message.chat.id
    bot.send_message(chat_id, "⏳ Sedang merangkum data dan menyusun Dashboard Excel...")

    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    
    cursor.execute("SELECT tanggal, jenis, kategori, nominal, keterangan FROM transaksi ORDER BY tanggal DESC")
    rows = cursor.fetchall()
    
    if not rows:
        bot.send_message(chat_id, "Belum ada data keuangan yang dicatat.")
        conn.close()
        return

    cursor.execute("SELECT SUM(nominal) FROM transaksi WHERE jenis='Pemasukan'")
    total_masuk = cursor.fetchone()[0] or 0
    cursor.execute("SELECT SUM(nominal) FROM transaksi WHERE jenis='Pengeluaran'")
    total_keluar = cursor.fetchone()[0] or 0
    saldo = total_masuk - total_keluar
    conn.close()

    wb = openpyxl.Workbook()
    
    # 1. Sheet Ringkasan (Dashboard)
    ws_dash = wb.active
    ws_dash.title = "Dashboard & Ringkasan"
    ws_dash.views.sheetView[0].showGridLines = True
    
    navy_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    accent_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    white_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Arial", size=11, bold=True)
    regular_font = Font(name="Arial", size=11)
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    ws_dash["A1"] = "KPI KEUANGAN PRIBADI"
    ws_dash["A1"].font = Font(name="Arial", size=14, bold=True, color="1F4E78")
    
    headers_dash = ["Indikator", "Total Nilai"]
    for col_num, header in enumerate(headers_dash, 1):
        cell = ws_dash.cell(row=3, column=col_num, value=header)
        cell.fill = navy_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center")

    ws_dash["A4"] = "Total Pemasukan"
    ws_dash["B4"] = "=SUMIF('Log Transaksi'!B:B, \"Pemasukan\", 'Log Transaksi'!D:D)"
    ws_dash["A5"] = "Total Pengeluaran"
    ws_dash["B5"] = "=SUMIF('Log Transaksi'!B:B, \"Pengeluaran\", 'Log Transaksi'!D:D)"
    ws_dash["A6"] = "Saldo Bersih (Net)"
    ws_dash["B6"] = "=B4-B5"
    
    for r in range(4, 7):
        ws_dash[f"A{r}"].font = regular_font
        ws_dash[f"B{r}"].font = bold_font if r==6 else regular_font
        ws_dash[f"B{r}"].number_format = '"Rp"#,##0'
        ws_dash[f"A{r}"].border = thin_border
        ws_dash[f"B{r}"].border = thin_border
    ws_dash["A6"].fill = accent_fill
    ws_dash["B6"].fill = accent_fill

    ws_dash["A9"] = "DISTRIBUSI ALOKASI PENGELUARAN"
    ws_dash["A9"].font = Font(name="Arial", size=12, bold=True, color="1F4E78")
    
    headers_kat = ["Kategori", "Total Pengeluaran", "Persentase (%)"]
    for col_num, header in enumerate(headers_kat, 1):
        cell = ws_dash.cell(row=11, column=col_num, value=header)
        cell.fill = navy_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center")
        
    kategori_list = ["Makanan", "Transportasi", "Kontrakan", "Belanja", "Hiburan", "Lainnya"]
    for idx, kat in enumerate(kategori_list, 12):
        ws_dash[f"A{idx}"] = kat
        ws_dash[f"B{idx}"] = f"=SUMIFS('Log Transaksi'!D:D, 'Log Transaksi'!B:B, \"Pengeluaran\", 'Log Transaksi'!C:C, \"{kat}\")"
        ws_dash[f"C{idx}"] = f"=IF(B5=0, 0, B{idx}/B5)" 
        
        ws_dash[f"A{idx}"].font = regular_font
        ws_dash[f"B{idx}"].font = regular_font
        ws_dash[f"C{idx}"].font = regular_font
        ws_dash[f"B{idx}"].number_format = '"Rp"#,##0'
        ws_dash[f"C{idx}"].number_format = '0.0%'
        ws_dash[f"A{idx}"].border = thin_border
        ws_dash[f"B{idx}"].border = thin_border
        ws_dash[f"C{idx}"].border = thin_border

    last_row = 12 + len(kategori_list)
    ws_dash[f"A{last_row}"] = "Total"
    ws_dash[f"B{last_row}"] = f"=SUM(B12:B{last_row-1})"
    ws_dash[f"C{last_row}"] = f"=SUM(C12:C{last_row-1})"
    
    for col in ["A", "B", "C"]:
        ws_dash[f"{col}{last_row}"].font = bold_font
        ws_dash[f"{col}{last_row}"].fill = accent_fill
        ws_dash[f"{col}{last_row}"].border = thin_border
    ws_dash[f"B{last_row}"].number_format = '"Rp"#,##0'
    ws_dash[f"C{last_row}"].number_format = '0.0%'

    # 2. Sheet Log Transaksi
    ws_log = wb.create_sheet(title="Log Transaksi")
    ws_log.views.sheetView[0].showGridLines = True
    
    headers_log = ["Tanggal & Waktu", "Jenis", "Kategori", "Nominal", "Keterangan"]
    for col_num, header in enumerate(headers_log, 1):
        cell = ws_log.cell(row=1, column=col_num, value=header)
        cell.fill = navy_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center")

    for row_num, row_data in enumerate(rows, 2):
        for col_num, val in enumerate(row_data, 1):
            cell = ws_log.cell(row=row_num, column=col_num, value=val)
            cell.font = regular_font
            cell.border = thin_border
            if col_num == 4:
                cell.number_format = '"Rp"#,##0'
                cell.alignment = Alignment(horizontal="right")

    for ws in [ws_dash, ws_log]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    filename = f"Laporan_Keuangan_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)

    kesimpulan = (
        f"📊 **ANALISIS AKUNTAN PRIBADI ANDA**\n"
        f"––––––––––––––––––––––––\n"
        f"📥 Total Pemasukan:  Rp {total_masuk:,.0f}\n"
        f"📤 Total Pengeluaran: Rp {total_keluar:,.0f}\n"
        f"––––––––––––––––––––––––\n"
        f"💰 **Saldo Bersih Anda:** Rp {saldo:,.0f}\n\n"
        f"💡 *Catatan:* "
    )
    if saldo < 0:
        kesimpulan += "Pengeluaran Anda membengkak melebihi pemasukan."
    elif total_keluar > (total_masuk * 0.7):
        kesimpulan += "Pengeluaran Anda sudah melewati 70% dari pemasukan."
    else:
        kesimpulan += "Kondisi keuangan sehat."

    bot.send_message(chat_id, kesimpulan, parse_mode="Markdown")
    
    with open(filename, 'rb') as file:
        bot.send_document(chat_id, file, caption="Berikut laporan keuangan terperinci Anda.")
        
    os.remove(filename)

# Run Bot
print("Akuntan Pribadi siap bekerja...")
bot.polling(none_stop=True)